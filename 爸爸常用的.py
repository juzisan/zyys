# -*- coding: utf-8 -*-
#!/usr/bin/env python
"""
Created on Sat Nov  5 10:50:13 2016

@author: hello
"""

import sys
import os
import re
import codecs
import copy
import math
import urllib
import time,datetime
import operator
import zipfile
import shutil
import glob
import pandas as pd
from multiprocessing import Pool as mpp
from multiprocessing.dummy import Pool as tpp
from math import log1p as ln
from matplotlib import pyplot
from pandas.tseries.offsets import *
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment,Border,Side,Font



def duo(shangdir):
    #文件名
    xlsx_nam=somedir+shangdir
    biao_or=pd.read_excel(xlsx_nam)#读excel
    biao_or=biao_or.sort_values(by =['总分'], ascending=False)
    biao_or.index = range(1,len(biao_or)+1)#重建索引
    biaotou = biao_or.columns.values.tolist()#pandas列名转list

    tizhi={}
    for i in range(len(biaotou)):#目的是删除多余的列。生成每题分值dict
        if biaotou[i].find(' - 得分') > 0:
            tim = biaotou[i][:biaotou[i].find(' - 得分')]#.replace(r'·','')
            tizhi[tim] = int(re.search(r'共(\d+)分', biaotou[i]).group(1))
            biaotou[i] = tim
        elif biaotou[i].find(' - 作答') > 0:
            biaotou[i] = '不用'
        elif biaotou[i] not in ['姓名','总分']:
            biaotou[i] = '不用'

    biao_or.columns =biaotou
    biaotou=[ x for x in biaotou if x != '不用']
    biao_or=biao_or[biaotou]#去掉无用列

    biao_or["合计"] = biao_or["总分"] - 150#处理扣总分
    del biao_or['总分']#删除总分列
    biaotou = biao_or.columns.values.tolist()

    for i in tizhi:#变成负分
        biao_or[i]=biao_or[i] - tizhi[i]
        biao_or[i]=biao_or[i].replace(0, np.nan)

    mean_mean =[ biao_or[x].mean() for x in biaotou[1:]]#pandas的mean出错太多了
    mean_mean.insert(0,'平均扣分')
    mean_mean = [ [x]  for x in mean_mean]
    mean_mean = pd.DataFrame(dict(zip(biaotou, mean_mean)))

    tj=pd.DataFrame([biao_or.count()])
    tj=tj.replace(0, np.nan)
    tj['姓名']=['扣分人数']#添加姓名列，内容是后面的两项，姓名列改内容
    tj=tj[biaotou]#按照biao_or.columns的顺序排列tj表，不排序就出错了，最后两行统计扣分人数和平均扣分
    biao_or=biao_or.append(tj)#把统计加入写入表

    biao_or = biao_or.append(mean_mean)#把平均数写入表，这次更新后出了一个bug，dataframe进行append后浮点数会变成复数，所以需要取real部分,根本问题是mean返回的是numpy的数组，是复数

    #print (biao_or.tail())
    print ('duo done {}'.format(shangdir))
    return biao_or


def main():
    #程序开始
    wb = Workbook()#建立excel文件
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))#创建边框对象
    for file_name in names:#开始循环处理文件夹下所有excel文件
        print (file_name)
        ws = wb.create_sheet(file_name.split('.')[0])#创建工作簿名
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE#页面方向
        ws.page_setup.paperSize = ws.PAPERSIZE_A4#纸张大小
        ws.page_margins.left = 0.25
        ws.page_margins.rigt = 0
        ws.page_margins.top = 0.2
        ws.page_margins.bottom = 0.2
        ws.page_margins.header = 0
        ws.page_margins.footer = 0#设置页边距
        ws.print_title_rows = '1:2'#设置打印标题为2行

        #print (dffff.tail())
        rows = dataframe_to_rows(duo(file_name))#pandas的dataframe转openpyxl的格式
        for row in rows:#不知道为啥有空行。openpyxl的序号从1开始。delete_rows()
            if len(row) >2:
                ws.append(row)
        ws.column_dimensions['A'].width = 3#设置列宽
        ws.column_dimensions['B'].width = 9
        for i in range(3,ws.max_column):#最大列数目，是整数类型
            ws.column_dimensions[get_column_letter(i)].width = 5
        ws.column_dimensions[get_column_letter(ws.max_column)].width = 6
        for i in list(ws.rows)[0]:#生成器转list才能迭代，第一行
            i.alignment = Alignment(wrap_text=True)#设置文本自动换行
        for row in ws.rows:
            for cell in row:
                cell.border =thin_border#设置边框

        ws.insert_rows(1)#插入标题行
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)#合并标题行
        p_title = file_name.split('.')[0]
        ws['A1'] = p_title
        ws['A1'].alignment = Alignment(horizontal='center')#设置标题居中
        ws['A1'].font = Font(size=20)#设置字号大小

    std=wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(std)#删除默认sheet，保存文件
    wb.save(somedir[:-1]+'6个班.xlsx')#excel写入sheat


#数据区
somedir=os.path.split(os.path.realpath(__file__))[0]+'\\试卷分析\\2017-18下学年新抚区毕业生五模\\'
names = os.listdir(somedir)
#数据区

if __name__ == "__main__":
    main()
    print ('all ok')

