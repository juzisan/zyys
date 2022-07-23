"""Created on Sat Oct 19 19:52:49 2019.

@author: Mrlily
整理爸爸的试卷分析
如果题号有错误，命名方式为    【客 | 一.1】
如果错误太多，再想办法重命名题号
一、生成负分表；
1.表头是班级，每班一个表
2.第一列是序号，按总分成绩排名
3.第二列是姓名
4.每题一列，得分转扣分，没扣分不显示，最后一列是合计
5.最后两列分别是平均扣分（扣分总和除以人数）和扣分人数
6.sheet重命名为班级
二、生成汇总表；
1.表头，合并表格并居中，字体增大
2.第一列为班级名，之后是每题扣分人数和平均扣分值
3.用pandas透视表把行转列
4.排列顺序安装班级名
5.横版页面布局，分两页
6.表格加边框
三、数据分析；
1.对班级统计
班级	    人数    总分    平均分    及格人数    及格率    优秀人数    优秀率    过差人数    过差率
2.对每题统计
班级    题号    本题满分    满分人数    满分率    及格人数    及格率    0分人数    0分率    本题平均分    本题得分率    最高分    最低分
根据模板生成每班的word试卷分析
"""

import locale
import os
import re
import shutil
import sys
import time

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from win32com.client import Dispatch

locale.setlocale(locale.LC_COLLATE, 'zh_CN.UTF8')  # 设置本地语言比较习惯为中文

p_title = ''


def timer(func):
    """计时器."""

    def warpper():
        """计时器内部."""
        print('\033[1;32;40mstart\033[0m')
        time1 = time.time()
        func()
        seconds = time.time() - time1
        m, s = divmod(seconds, 60)
        print("\033[1;32;40mthe run time is %02d:%.6f\033[0m" % (m, s))

    return warpper


def save_file(file_str, data_df):
    """重复部分."""
    file_str = file_str + '班'
    # 文件名
    wb = Workbook()  # 建立excel文件
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))  # 创建边框对象
    ws = wb.create_sheet(file_str)  # 创建工作簿名
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 页面方向
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0
    ws.page_margins.footer = 0  # 设置页边距
    ws.print_title_rows = '1:2'  # 设置打印标题为2行
    rows = dataframe_to_rows(data_df)  # pandas 的dataframe 转openpyxl 的格式
    for row in rows:  # 不知道为啥有空行。openpyxl的序号从1开始。delete_rows()
        if len(row) > 2:
            ws.append(row)
    ws.column_dimensions['A'].width = 3  # 设置列宽
    ws.column_dimensions['B'].width = 9
    for i in range(3, 12):  # 最大列数目，是整数类型
        ws.column_dimensions[get_column_letter(i)].width = 4
    for i in range(12, ws.max_column):  # 最大列数目，是整数类型
        ws.column_dimensions[get_column_letter(i)].width = 5
    ws.column_dimensions[get_column_letter(ws.max_column)].width = 6
    for i in list(ws.rows)[0]:  # 生成器转list才能迭代，第一行
        i.alignment = Alignment(wrap_text=True)  # 设置文本自动换行
    for row in ws.rows:
        for cell in row:
            cell.border = thin_border  # 设置边框

    ws.insert_rows(1)  # 插入标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    # 合并标题行
    ws['A1'] = file_str
    ws['A1'].alignment = Alignment(horizontal='center')  # 设置标题居中
    ws['A1'].font = Font(size=20)  # 设置字号大小

    del wb['Sheet']  # 删除默认sheet，保存文件
    wb.save('试卷分析' + file_str + '.xlsx')  # excel写入sheet
    del wb, ws


def save_file2(file_str, data_df):
    """写入Excel文件."""
    wb = Workbook()  # 创建workbook实例
    ws = wb.active
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))  # 创建边框对象
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 页面方向
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0
    ws.page_margins.footer = 0  # 设置页边距
    ws.print_title_cols = 'A:A'

    for row in ws.rows:
        for cell in row:
            cell.border = thin_border  # 设置边框
    ws.insert_rows(0)  # 插入标题行

    for r in dataframe_to_rows(data_df, index=True, header=True):
        ws.append(r)

    ws.column_dimensions['A'].width = 4  # 设置列宽
    for i in range(2, ws.max_column + 1):  # 最大列数目，是整数类型
        ws.column_dimensions[get_column_letter(i)].width = 5
    for i in range(3, ws.max_row + 1):  # 最大行数目，是整数类型
        ws.row_dimensions[i].height = 43

    for i in range(2, ws.max_column, 2):  # 合并表头
        end_i = i + 1
        ws.merge_cells(start_row=2, start_column=i,
                       end_row=2, end_column=end_i)

    for row in list(ws.rows)[1:]:
        for cell in row:
            cell.border = thin_border  # 设置边框

    ziti = ws['A2:' + get_column_letter(ws.max_column) + str(ws.max_row)]
    for row in list(ziti):  # 单元格选择区域是tuple，转换成list就能赋值了
        for cell in row:
            cell.font = Font(size=11)
            cell.alignment = Alignment(wrap_text=True)  # 设置文本自动换行

    ws['A3'] = ws['A4'].value  # 单元格值需要value
    ws.delete_rows(4)  # 删行
    ws.merge_cells('D1:U1')
    ws['D1'] = '数学' + file_str  # '七年_________考试教学质量分析统计表（数学）'改成前面的了
    ws['D1'].alignment = Alignment(horizontal='center')  # 设置标题居中
    ws['D1'].font = Font(size=20)  # 设置字号大小

    wb.save('汇总.xlsx')  # excel写入sheet
    os.startfile('汇总.xlsx')
    del wb, ws


def save_file3(data_1, data_2):
    """写入Excel文件."""
    global p_title
    wb = Workbook()  # 创建workbook实例
    ws = wb.active
    # dataframe转行
    for row in dataframe_to_rows(data_1, index=True, header=True):
        ws.append(row)
    for row in dataframe_to_rows(data_2, index=True, header=True):
        ws.append(row)
    rows_num = ws.max_row

    for i in range(1, rows_num + 1):
        rows_list = [cell.value is None for cell in ws[i]]
        if all(rows_list):
            ws.delete_rows(i)  # 删除行，注意会不会删错
    ws.delete_cols(1)  # 删除列
    wb.save('描述.xlsx')  # excel写入sheet
    os.startfile('描述.xlsx')
    del wb, ws

    class_names_set = data_1['班级'].unique()  # 班级的set
    table_1_df = data_1.groupby(['班级'])
    table_2_df = data_2.groupby(['班级'])
    print(class_names_set)
    dir_str = os.getcwd()  # 本地目录
    src_file = os.path.join(dir_str, '试卷分析模板.doc')

    msword = Dispatch('Word.Application')  # 打开微软word程序，后面得关闭
    msword.Visible = 1  # 显示word窗口
    for name_of_class in class_names_set:
        file_name = name_of_class + '试卷分析.doc'
        dst_file = os.path.join(dir_str, file_name)
        if os.path.exists(dst_file):
            os.unlink(dst_file)  # 文件存在则删除文件，不然文件存在的时候，复制会出错
        shutil.copyfile(src_file, dst_file)  # 复制文件
        print(f'{dst_file = }')
        doc = msword.Documents.Open(dst_file)  # 打开word文件
        p1 = doc.Paragraphs(1)
        replace_t = ('xxx-xxx', False, False, False, False, False, True, 1, True, name_of_class + p_title, 2)
        p1.Range.Find.Execute(*replace_t)
        # 第一段文本替换，参数用*解包
        table_s = doc.Tables  # word文件中所有表格的集合
        num_table = doc.Tables.Count  # word文件中有几个表格，序号从1开始，不是0
        print(f'{num_table = }\n')
        # 一、抽样成绩统计
        num_columns = table_s(1).Range.Columns.Count  # 表格的列数
        # print(f'{num_rows = }', f'{num_columns = }')
        for num_col in range(1, num_columns + 1):  # 只有一行，所以行数为1
            table_s(1).Cell(2, num_col).Range.Text = str(
                table_1_df.get_group(name_of_class).iloc[0, num_col - 1])
        # 列对应的dataframe表，第一行是0，不算索引第一列是0

        # 二、逐题分析
        num_rows = table_s(2).Range.Rows.Count  # 表格的行数
        num_columns = table_s(2).Range.Columns.Count
        # print(f'{num_rows = }', f'{num_columns = }')
        for num_row in range(3, num_rows + 1):
            for num_col in range(1, num_columns + 1):
                table_s(2).Cell(num_row, num_col).Range.Text = str(
                    table_2_df.get_group(name_of_class).iloc[num_row - 3, num_col])
                # word的行从第3行开始，dataframe第1列是班级，舍弃，所以列从1开始
        doc.Save()
        doc.Close()  # 保存并退出
    msword.Quit()  # 手动关闭word程序


def tiaozh(num_df):
    """."""
    num_df = num_df.dropna(axis=1, how='all')  # 最后一行，有数据的是每题分值
    num_df = num_df.T  # 行转列
    num_df = num_df.reset_index()
    num_df.columns = ["题号备份", "满分值"]
    num_df["满分值"].replace(
        r'^得分/共(\d+)分$', r'\1', regex=True, inplace=True)  # 替换
    num_df["满分值"] = num_df["满分值"].astype('int', copy=False)
    num_df['题号'] = num_df['题号备份'].replace(
        regex=True, inplace=False, to_replace=r'^(主|客) \| ', value=r'')  # 用正则表达式替换，注意inplace
    num_df[['大题', '小题']] = num_df['题号'].str.split('.', n=1, expand=True)
    if_na = num_df[num_df['小题'].isna()].index.tolist()

    if if_na:
        print(f'题号有错误:  {if_na}\n退出程序')
        sys.exit()
    else:
        print('题号无错误')

    # num_df['题号'] = num_df['大题'] + '.' + num_df['小题']
    big_set = num_df['大题'].unique()
    rename_dict = num_df.set_index(['题号备份'])["题号"].to_dict()
    score_dict = num_df.set_index(['题号'])["满分值"].to_dict()
    big_num_list = []
    big_name_list = []
    for i in big_set:
        big_num_list.append(
            [i, num_df[num_df['大题'] == i].sum()['满分值']])
        big_name_list.append(
            [i, num_df['题号'][num_df['大题'] == i].tolist()])
    # print(f'{num_df = }')
    # print(num_df.dtypes)
    # print(big_num_list, big_name_list, rename_dict, score_dict, sep='\n\n')
    return rename_dict, score_dict, big_num_list, big_name_list


def chuli(file_str):
    """共有."""
    global p_title
    p_title = re.findall('^小分表 - ([\\w\\W]*).xlsx$', file_str)[0]
    print(f'考试标题： {p_title}')
    original_df = pd.read_excel(io=file_str, sheet_name=0, dtype={'班级': 'str'}, skiprows=1)
    # read_excel必须要sheet_name，0为第一个，None为所有都读取，读到一个字典里
    original_df = original_df.sort_values(by=['总分'], ascending=False)
    original_df.index = range(1, len(original_df) + 1)  # 重建索引 从1开始

    rename_dict, score_dict, big_num_list, big_name_list = tiaozh(
        original_df.tail(1).copy(deep=True))

    # print(rename_dict, score_dict, big_num_list, big_name_list)
    sum_score_number = sum((j for i, j in big_num_list))  # 求总分
    print(f"\n总分：   {sum_score_number}  ")  # 核对总分

    # original_df = pd.concat([original_df, score_dataframe], ignore_index=True) # 合并表格
    original_df.drop(original_df.tail(1).index, inplace=True)
    # 删除最后一行，之前的只是另存
    original_df.rename(columns=rename_dict, inplace=True)
    # 把列重命名,tiaozh
    rename_lst = ['姓名', '班级', '总分'] + list(rename_dict.values())
    original_df = original_df[rename_lst]
    # 去掉无用列,tiaozh
    print('整理后的dataframe：\n')
    # print(original_df)

    """共有."""
    """
    1.生成负分表；
    2.生成汇总表；
    3.数据分析
    """
    original_df["合计"] = original_df["总分"].astype('float') - sum_score_number  # 处理扣总分
    """复制到分析表"""
    analyze_df = original_df.copy(deep=True)
    """复制到分析表"""
    del original_df['总分']  # 删除总分列

    for j in score_dict:  # 变成负分.astype('float')
        original_df[j] = original_df[j].astype('float') - score_dict.get(j)
        original_df[j] = original_df[j].replace(0, np.nan)

    original_df = original_df.groupby(['班级'])
    # 按照班级分类,分类并不重建索引，沿用没分类之前的索引
    # print(original_df)
    summary_list = []
    for key, value in original_df:
        mean_series = value.mean(numeric_only=True)  # 平均扣分 行
        mean_series['姓名'] = "平均扣分"  # 重新赋值
        count_series = value.count()  # 扣分人数 行
        count_series['姓名'] = "扣分人数"  # 重新赋值
        total_df = pd.concat([mean_series, count_series], axis=1).T
        value = pd.concat([value, total_df])
        # value = value.append(mean_series, ignore_index=True)  # 插入平均扣分行
        # value = value.append(count_series, ignore_index=True)  # 插入扣分人数行
        value['班级'] = str(key) + '班'
        summary_list.append(value[-2:].copy(deep=True))

        del value['班级']  # 删除 班级 列
        value.index = range(1, len(value) + 1)  # 重建索引
        """1.负分表保存文件"""
        save_file(key, value)

    summary_df = pd.concat(
        summary_list, ignore_index=True)  # 合并dataframe
    headline_list = summary_df.columns.values.tolist()  # 把列名转换成 list 格式
    headline_list = headline_list[2:-1]  # 删除前面和后面的列名
    headline_new_list = []
    for i in headline_list:
        headline_new_list.append((i, '扣分人数'))
        headline_new_list.append((i, '平均扣分'))
    # 制作复合列名
    summary_df = summary_df.pivot_table(index='班级', columns='姓名', margins=False)
    # 创建透视表，因为列名需要占用两行
    summary_df = summary_df[headline_new_list]  # 重现对列名进行格式化
    '''2.汇总表保存文件'''
    save_file2(p_title, summary_df)
    print('汇总表：\n')
    # print(summary_df)

    '''3.数据分析'''
    '''总人数count()后面不用[]，有筛选的就要用[]了'''
    big_keep_list = []
    print('汇总题号：\n')
    for b_key, b_value in big_name_list:  # 返回值来自于tiaozh
        # print(b_key, b_value)
        analyze_df[b_key] = analyze_df[b_value].sum(
            axis=1)
        # 根据列名求和,来自于tiaozh
        big_keep_list.append(b_key)

    analyze_df = analyze_df[['班级', '总分'] + big_keep_list]  # 去掉无用列

    # analyze_1_dataframe = analyze_df.copy(deep=True)
    analyze_list = [['全年级', analyze_df.copy(deep=True)]]  # 复制数据表

    analyze_df = analyze_df.groupby(['班级'])  # 按照班级分类
    for key, value in analyze_df:
        analyze_list.append([str(key) + '班', value])

    class_summary_lst = []
    class_lst = []

    for key, value in analyze_list:
        value.index = range(1, len(value) + 1)  # 重建索引
        total_people = value['总分'].count()
        series_1 = pd.Series(key, ['班级'])
        series_1['人数'] = total_people
        series_1['总分'] = sum_score_number
        series_1['平均分'] = round(value["总分"].mean(), 1)
        series_1['及格人数'] = value[value["总分"] >
                                 sum_score_number * 0.59].count()["总分"]
        series_1['及格率'] = round(series_1['及格人数'] / total_people * 100, 1)
        series_1['优秀人数'] = value[value["总分"] >
                                 sum_score_number * 0.84].count()["总分"]
        series_1['优秀率'] = round(series_1['优秀人数'] / total_people * 100, 1)
        series_1['过差人数'] = value[value["总分"] <
                                 sum_score_number * 0.40].count()["总分"]
        series_1['过差率'] = round(series_1['过差人数'] / total_people * 100, 1)
        class_summary_lst.append(series_1)

        for problem_num, score_num in big_num_list:
            series_2 = pd.Series(key, ['班级'])
            series_2['题号'] = problem_num
            series_2['本题满分'] = score_num
            series_2['满分人数'] = value[value[problem_num]
                                     > score_num * 0.997].count()[problem_num]
            series_2['满分率'] = round(series_2['满分人数'] / total_people * 100, 1)
            series_2['及格人数'] = value[value[problem_num]
                                     > score_num * 0.597].count()[problem_num]
            series_2['及格率'] = round(series_2['及格人数'] / total_people * 100, 1)
            series_2['0分人数'] = value[value[problem_num] == 0].count()[
                problem_num]
            series_2['0分率'] = round(series_2['0分人数'] / total_people * 100, 1)
            series_2['本题平均分'] = round(value[problem_num].mean(), 1)
            series_2['本题得分率'] = round(
                value[problem_num].sum() / total_people / score_num * 100, 1)
            series_2['最高分'] = value[problem_num].max()
            series_2['最低分'] = value[problem_num].min()
            class_lst.append(series_2)
    class_s_df = pd.concat(class_summary_lst, axis=1)
    class_s_df = class_s_df.T
    class_df = pd.concat(class_lst, axis=1)
    class_df = class_df.T
    # series转dataframe，倒置后和Excel一样
    class_s_df = class_s_df.astype(
        {'人数': 'int', '总分': 'int', '及格人数': 'int', '过差人数': 'int'})
    # 把数据类型改成整数，不然是浮点数
    class_df = class_df.astype(
        {'本题满分': 'int', '满分人数': 'int', '及格人数': 'int', '0分人数': 'int'})
    # 把数据类型改成整数，不然是浮点数
    print('描述dataframe：\n')
    # print(f'{class_s_df}\n\n{class_df}')
    save_file3(class_s_df, class_df)

    print('OK')


# 数据区


# 数据区

@timer
def main():
    """程序开始."""
    names = os.listdir(os.path.split(os.path.realpath(__file__))[0])
    names = [i for i in names if re.match('小分表.*.xlsx', i)]
    if names:
        if len(names) > 1:
            print('请输入序号：')
            for i, value in enumerate(names):
                print(i, '代表：  ', value)
            select_num = int(input("输入转换的序号："))
            chuli(names[select_num])
        else:
            chuli(names[0])
    else:
        print('缺少文件')


if __name__ == "__main__":
    main()
    print('all ok')

'''
'''