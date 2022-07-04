"""Created on Sat Oct 19 19:52:49 2019.

@author: Mrlily
整理爸爸的试卷分析
如果题号有错误，命名方式为    【客 | 一.1】
如果错误太多，再想办法重命名题号
一、生成负分表；
1.表头是班级，每班一个表
2.第一列是序号，按总分成绩排名
3.第二列是姓名
4.每题一列，得分转扣分，没扣分不显示，最后一列是合计
5.最后两列分别是平均扣分（扣分总和除以人数）和扣分人数
6.sheet重命名为班级
二、生成汇总表；
1.表头，合并表格并居中，字体增大
2.第一列为班级名，之后是每题扣分人数和平均扣分值
3.用pandas透视表把行转列
4.排列顺序安装班级名
5.横版页面布局，分两页
6.表格加边框
三、数据分析；
1.对班级统计
班级	    人数    总分    平均分    及格人数    及格率    优秀人数    优秀率    过差人数    过差率
2.对每题统计
班级    题号    本题满分    满分人数    满分率    及格人数    及格率    0分人数    0分率    本题平均分    本题得分率    最高分    最低分
"""

import os
import re
import sys
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side, Font
import time
import locale


locale.setlocale(locale.LC_COLLATE, 'zh_CN.UTF8')  # 设置本地语言比较习惯为中文


def timer(func):
    """ 计时器. """

    def warpper():
        """计时器内部."""
        print('\033[1;32;40mstart\033[0m')
        time1 = time.time()
        func()
        seconds = time.time() - time1
        m, s = divmod(seconds, 60)
        print("\033[1;32;40mthe run time is %02d:%.6f\033[0m" % (m, s))

    return warpper


def save_file(filename_str, data_dataframe):
    """重复部分."""
    filename_str = filename_str + '班'
    # 文件名

    wb = Workbook()  # 建立excel文件
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))  # 创建边框对象

    ws = wb.create_sheet(filename_str)  # 创建工作簿名
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 页面方向
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0
    ws.page_margins.footer = 0  # 设置页边距
    ws.print_title_rows = '1:2'  # 设置打印标题为2行
    rows = dataframe_to_rows(data_dataframe)  # pandas 的dataframe 转openpyxl 的格式
    for row in rows:  # 不知道为啥有空行。openpyxl的序号从1开始。delete_rows()
        if len(row) > 2:
            ws.append(row)
    ws.column_dimensions['A'].width = 3  # 设置列宽
    ws.column_dimensions['B'].width = 9
    for i in range(3, 12):  # 最大列数目，是整数类型
        ws.column_dimensions[get_column_letter(i)].width = 4
    for i in range(12, ws.max_column):  # 最大列数目，是整数类型
        ws.column_dimensions[get_column_letter(i)].width = 5
    ws.column_dimensions[get_column_letter(ws.max_column)].width = 6
    for i in list(ws.rows)[0]:  # 生成器转list才能迭代，第一行
        i.alignment = Alignment(wrap_text=True)  # 设置文本自动换行
    for row in ws.rows:
        for cell in row:
            cell.border = thin_border  # 设置边框

    ws.insert_rows(1)  # 插入标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=ws.max_column)  # 合并标题行
    p_title = filename_str
    ws['A1'] = p_title
    ws['A1'].alignment = Alignment(horizontal='center')  # 设置标题居中
    ws['A1'].font = Font(size=20)  # 设置字号大小

    del wb['Sheet']  # 删除默认sheet，保存文件
    wb.save('试卷分析' + filename_str + '.xlsx')  # excel写入sheet
    del wb, ws


def save_file2(filename_str, data_dataframe):
    wb = Workbook()  # 创建workbook实例
    ws = wb.active
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))  # 创建边框对象
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 页面方向
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0
    ws.page_margins.footer = 0  # 设置页边距
    ws.print_title_cols = 'A:A'

    for row in ws.rows:
        for cell in row:
            cell.border = thin_border  # 设置边框
    ws.insert_rows(0)  # 插入标题行

    for r in dataframe_to_rows(data_dataframe, index=True, header=True):
        ws.append(r)

    ws.column_dimensions['A'].width = 4  # 设置列宽
    for i in range(2, ws.max_column + 1):  # 最大列数目，是整数类型
        ws.column_dimensions[get_column_letter(i)].width = 5
    for i in range(3, ws.max_row + 1):  # 最大行数目，是整数类型
        ws.row_dimensions[i].height = 43

    for i in range(2, ws.max_column, 2):  # 合并表头
        end_i = i + 1
        ws.merge_cells(start_row=2, start_column=i,
                       end_row=2, end_column=end_i)

    for row in list(ws.rows)[1:]:
        for cell in row:
            cell.border = thin_border  # 设置边框

    ziti = ws['A2:' + get_column_letter(ws.max_column) + str(ws.max_row)]
    for row in list(ziti):  # 单元格选择区域是tuple，转换成list就能赋值了
        for cell in row:
            cell.font = Font(size=11)
            cell.alignment = Alignment(wrap_text=True)  # 设置文本自动换行

    ws['A3'] = ws['A4'].value  # 单元格值需要value
    ws.delete_rows(4)  # 删行
    ws.merge_cells('D1:U1')
    p_title = '数学' + filename_str  # '七年_________考试教学质量分析统计表（数学）'改成前面的了
    ws['D1'] = p_title
    ws['D1'].alignment = Alignment(horizontal='center')  # 设置标题居中
    ws['D1'].font = Font(size=20)  # 设置字号大小

    wb.save('汇总.xlsx')  # excel写入sheet
    os.startfile('汇总.xlsx')
    del wb, ws


def save_file3(*data_list):
    wb = Workbook()  # 创建workbook实例
    ws = wb.active

    for data_df in data_list:
        for row in dataframe_to_rows(data_df, index=True, header=True):
            ws.append(row)
    rows_num = ws.max_row

    for i in range(1, rows_num + 1):
        rows_list = [cell.value is None for cell in ws[i]]
        if all(rows_list):
            ws.delete_rows(i)  # 删除行，注意会不会删错
    ws.delete_cols(1)  # 删除列
    wb.save('描述.xlsx')  # excel写入sheet
    os.startfile('描述.xlsx')
    del wb, ws


def tiaozh(num_dataframe):
    num_dataframe = num_dataframe.dropna(axis=1, how='all')  # 最后一行，有数据的是每题分值
    num_dataframe = num_dataframe.T  # 行转列
    num_dataframe = num_dataframe.reset_index()
    num_dataframe.columns = ["题号", "满分值"]
    num_dataframe["满分值"].replace(r'得分/共(\d+)分', r'\1', regex=True, inplace=True)  # 替换
    num_dataframe["满分值"] = num_dataframe["满分值"].astype('int', copy=False)
    num_dataframe['题号备份'] = num_dataframe['题号']
    # num_dataframe['主客观'], num_dataframe['题号'] = num_dataframe['题号'].str.split('|', 1).str
    num_dataframe[['主客观', '题号']] = num_dataframe['题号'].str.split('|', n=1, expand=True)
    num_dataframe['题号'].replace(regex=True, inplace=True, to_replace=r' ', value=r'')
    num_dataframe[['大题', '小题']] = num_dataframe['题号'].str.split('.', n=1, expand=True)
    if_na = num_dataframe[num_dataframe['小题'].isna()].index.tolist()
    
    if if_na:
        print(f'题号有错误:  {if_na}\n退出程序')
        sys.exit()
    else:
        print('题号无错误')

    # num_dataframe['题号'] = num_dataframe['大题'] + '.' + num_dataframe['小题']
    big_set = num_dataframe['大题'].unique()
    rename_dict = num_dataframe.set_index(['题号备份'])["题号"].to_dict()
    score_dict = num_dataframe.set_index(['题号'])["满分值"].to_dict()
    big_num_list = []
    big_name_list = []
    for i in big_set:
        big_num_list.append([i, num_dataframe[num_dataframe['大题'] == i].sum()['满分值']])
        big_name_list.append([i, num_dataframe['题号'][num_dataframe['大题'] == i].tolist()])
    # print(num_dataframe)
    # print(num_dataframe.dtypes)
    # print(big_num_list, big_name_list, rename_dict, score_dict, sep='\n\n')
    return rename_dict, score_dict, big_num_list, big_name_list 


def chuli(filename_str_chuli):
    """共有."""
    p_title = re.findall('^小分表 - ([\\w\\W]*).xlsx$', filename_str_chuli)[0]
    print(f'考试标题： {p_title}')
    original_dataframe = pd.read_excel(filename_str_chuli, skiprows=1, dtype={
        '班级': 'str', })  # 读 excel '总分':'int'
    original_dataframe = original_dataframe.sort_values(by=['总分'], ascending=False)
    original_dataframe.index = range(1, len(original_dataframe) + 1)  # 重建索引 从1开始
    
    rename_dict, score_dict, big_num_list, big_name_list = tiaozh(original_dataframe.tail(1).copy(deep=True))
    
    # print(rename_dict, score_dict, big_num_list, big_name_list)
    sum_score_number = sum([j for i, j in big_num_list])  # 求总分
    print(f"\n总分：   {sum_score_number}  ")  # 核对总分

    # original_dataframe = pd.concat([original_dataframe, score_dataframe], ignore_index=True) # 合并表格
    original_dataframe.drop(original_dataframe.tail(1).index, inplace=True)  # 删除最后一行，之前的只是另存
    original_dataframe.rename(columns=rename_dict, inplace=True)  # 把列重命名,tiaozh
    original_dataframe = original_dataframe[['姓名', '班级', '总分'] + list(rename_dict.values())]  # 去掉无用列,tiaozh
    print('整理后的dataframe：\n')
    # print(original_dataframe)

    """共有."""
    """
    1.生成负分表；
    2.生成汇总表；
    3.数据分析
    """
    original_dataframe["合计"] = original_dataframe["总分"].astype('float') - sum_score_number  # 处理扣总分
    """复制到分析表"""
    analyze_dataframe = original_dataframe.copy(deep=True)
    """复制到分析表"""
    del original_dataframe['总分']  # 删除总分列

    for j in score_dict:  # 变成负分.astype('float')
    
        original_dataframe[j] = original_dataframe[j].astype('float') - score_dict.get(j)
        original_dataframe[j] = original_dataframe[j].replace(0, np.nan)

    original_dataframe = original_dataframe.groupby(['班级'])  # 按照班级分类,分类并不重建索引，沿用没分类之前的索引
    # print(original_dataframe)
    summary_list = []
    for key, value in original_dataframe:
        mean_series = value.mean(numeric_only=True)  # 平均扣分 行
        mean_series['姓名'] = "平均扣分"  # 重新赋值

        count_series = value.count()  # 扣分人数 行
        count_series['姓名'] = "扣分人数"  # 重新赋值
        value = value.append(mean_series, ignore_index=True)  # 插入平均扣分行
        value = value.append(count_series, ignore_index=True)  # 插入扣分人数行
        value['班级'] = str(key) + '班'
        summary_list.append(value[-2:].copy(deep=True))

        del value['班级']  # 删除 班级 列
        value.index = range(1, len(value) + 1)  # 重建索引
        """1.负分表保存文件"""
        save_file(key, value)

    summary_dataframe = pd.concat(summary_list, ignore_index=True)  # 合并dataframe
    headline_list = summary_dataframe.columns.values.tolist()  # 把列名转换成 list 格式
    headline_list = headline_list[2:-1]  # 删除前面和后面的列名
    headline_new_list = []
    for i in headline_list:
        headline_new_list.append((i, '扣分人数'))
        headline_new_list.append((i, '平均扣分'))
    # 制作复合列名
    summary_dataframe = summary_dataframe.pivot_table(
        index='班级', columns='姓名', margins=False)  # 创建透视表，因为列名需要占用两行
    summary_dataframe = summary_dataframe[headline_new_list]  # 重现对列名进行格式化
    '''2.汇总表保存文件'''
    save_file2(p_title, summary_dataframe)
    print('汇总表：\n')
    # print(summary_dataframe)

    '''3.数据分析'''
    '''总人数count()后面不用[]，有筛选的就要用[]了'''
    big_keep_list = []
    print('汇总题号：\n')
    for bignumber_key, bignumber_value in big_name_list:  # 返回值来自于tiaozh
        # print(bignumber_key, bignumber_value)
        analyze_dataframe[bignumber_key] = analyze_dataframe[bignumber_value].sum(axis=1)  # 根据列名求和,来自于tiaozh
        big_keep_list.append(bignumber_key)

    analyze_dataframe = analyze_dataframe[['班级', '总分'] + big_keep_list]  # 去掉无用列

    # analyze_1_dataframe = analyze_dataframe.copy(deep=True)  
    analyze_list = [['全年级', analyze_dataframe.copy(deep=True)]]  # 复制数据表

    analyze_dataframe = analyze_dataframe.groupby(['班级'])  # 按照班级分类
    for key, value in analyze_dataframe:
        analyze_list.append([str(key) + '班', value])

    class_summary_dataframe = pd.DataFrame()
    class_dataframe = pd.DataFrame()

    for key, value in analyze_list:
        value.index = range(1, len(value) + 1)  # 重建索引
        total_people = value['总分'].count()
        series_1 = pd.Series(key, ['班级'])
        series_1['人数'] = total_people
        series_1['总分'] = sum_score_number
        series_1['平均分'] = round(value["总分"].mean(), 1)
        series_1['及格人数'] = value[value["总分"] > sum_score_number * 0.59].count()["总分"]
        series_1['及格率'] = round(series_1['及格人数'] / total_people * 100, 1)
        series_1['优秀人数'] = value[value["总分"] > sum_score_number * 0.84].count()["总分"]
        series_1['优秀率'] = round(series_1['优秀人数'] / total_people * 100, 1)
        series_1['过差人数'] = value[value["总分"] < sum_score_number * 0.40].count()["总分"]
        series_1['过差率'] = round(series_1['过差人数'] / total_people * 100, 1)
        class_summary_dataframe = class_summary_dataframe.append(series_1, ignore_index=True)

        for problem_num, score_num in big_num_list:
            series_2 = pd.Series(key, ['班级'])
            series_2['题号'] = problem_num
            series_2['本题满分'] = score_num
            series_2['满分人数'] = value[value[problem_num] > score_num * 0.997].count()[problem_num]
            series_2['满分率'] = round(series_2['满分人数'] / total_people * 100, 1)
            series_2['及格人数'] = value[value[problem_num] > score_num * 0.597].count()[problem_num]
            series_2['及格率'] = round(series_2['及格人数'] / total_people * 100, 1)
            series_2['0分人数'] = value[value[problem_num] == 0].count()[problem_num]
            series_2['0分率'] = round(series_2['0分人数'] / total_people * 100, 1)
            series_2['本题平均分'] = round(value[problem_num].mean(), 1)
            series_2['本题得分率'] = round(value[problem_num].sum() / total_people / score_num * 100, 1)
            series_2['最高分'] = value[problem_num].max()
            series_2['最低分'] = value[problem_num].min()
            class_dataframe = class_dataframe.append(series_2, ignore_index=True)
    print('描述dataframe：\n')
    # print(f'{class_summary_dataframe}\n\n{class_dataframe}')
    save_file3(class_summary_dataframe, class_dataframe)

    print('OK')


# 数据区


# 数据区

@timer
def main():
    """程序开始."""
    names = os.listdir(os.path.split(os.path.realpath(__file__))[0])
    names = [i for i in names if re.match('小分表.*.xlsx', i)]
    if names:
        if len(names) > 1:
            print('请输入序号：')
            for i, value in enumerate(names):
                print(i, '代表：  ', value)
            select_num = int(input("输入转换的序号："))
            chuli(names[select_num])
        else:
            chuli(names[0])
    else:
        print('缺少文件')


if __name__ == "__main__":
    main()
    print('all ok')

'''



'''
