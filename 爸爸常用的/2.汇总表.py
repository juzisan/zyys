"""Created on Sat Nov  5 10:50:13 2016.

@author: hello
"""


import numpy as np
import os
import pandas as pd
import re
import time
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def timer(func):
    """计时器."""
    def warpper():
        """计时器内部."""
        print('\033[1;32;40mstart\033[0m')
        time1 = time.time()
        func()
        seconds = time.time() - time1
        m, s = divmod(seconds, 60)
        print("\033[1;32;40mthe run time is %02d:%.6f\033[0m" % (m, s))
    return warpper


def duo(xlsx_nam):
    """重复."""
    biao_or = pd.read_excel("试卷分析" + xlsx_nam + ".xlsx",
                            index_col=0, skiprows=1)  # 读excel
    biao = biao_or[-2:].copy()
    biao['班级'] = xlsx_nam
    print('duo done {}'.format(xlsx_nam))
    return biao


@timer
def main():
    """程序开始."""
    names = os.listdir(os.path.split(os.path.realpath(__file__))[0])
    shanchu = [i for i in names if i.startswith('汇总表')]
    for i in shanchu:
        os.remove(i)
    names = os.listdir(os.path.split(os.path.realpath(__file__))[0])
    print(names)
    p_titlel = [i for i in names if re.match('^小分表 - ([\\w\\W]*).xlsx$', i)]
    p_title = re.findall('^小分表 - ([\\w\\W]*).xlsx$', p_titlel[0])[0]
    print(p_title)
    names = [i for i in names if i.count('试卷分析')]

    names1 = [i for i in names if re.match('^试卷分析\\d班.xlsx$', i)]
    names2 = [i for i in names if re.match('^试卷分析\\d{2}班.xlsx$', i)]
    # (高湾|新抚)
    names = names1 + names2
    names = [i.lstrip("试卷分析") for i in names]
    names = [i.rstrip(".xlsx") for i in names]
    data_hebing = map(duo, names)  # [duo(i) for i in names]
    data_hebing = pd.concat(data_hebing, ignore_index=True)

    data_hebing = data_hebing.replace(np.nan, 0)
    biaotou = data_hebing.columns.values.tolist()

    data_hebing = data_hebing.pivot_table(
        index='班级', columns='姓名',  margins=False)
    data_hebing = data_hebing.loc[names]  # 单双位排序

    biaotou = biaotou[1:-1]
    biaotou3 = []
    for i in biaotou:
        biaotou3.append((i, '扣分人数'))
        biaotou3.append((i, '平均扣分'))

    data_hebing = data_hebing[biaotou3]

    data_hebing.index.map(lambda x: str(x) + '班')
    print(data_hebing.index)

    wb = Workbook()  # 创建workbook实例
    ws = wb.active
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))  # 创建边框对象
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 页面方向
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 纸张大小
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0
    ws.page_margins.footer = 0  # 设置页边距
    ws.print_title_cols = 'A:A'

    for row in ws.rows:
        for cell in row:
            cell.border = thin_border  # 设置边框
    ws.insert_rows(0)  # 插入标题行

    for r in dataframe_to_rows(data_hebing, index=True, header=True):
        ws.append(r)

    ws.column_dimensions['A'].width = 4  # 设置列宽
    for i in range(2, ws.max_column+1):  # 最大列数目，是整数类型
        ws.column_dimensions[get_column_letter(i)].width = 5
    for i in range(3, ws.max_row+1):  # 最大行数目，是整数类型
        ws.row_dimensions[i].height = 43

    for i in range(2, ws.max_column, 2):  # 合并表头
        end_i = i + 1
        ws.merge_cells(start_row=2, start_column=i,
                       end_row=2, end_column=end_i)

    for row in list(ws.rows)[1:]:
        for cell in row:
            cell.border = thin_border  # 设置边框

    ziti = ws['A2:'+get_column_letter(ws.max_column)+str(ws.max_row)]
    for row in list(ziti):  # 单元格选择区域是tuple，转换成list就能赋值了
        for cell in row:
            cell.font = Font(size=11)
            cell.alignment = Alignment(wrap_text=True)  # 设置文本自动换行

    ws['A3'] = ws['A4'].value  # 单元格值需要value
    ws.delete_rows(4)  # 删行
    ws.merge_cells('D1:U1')
    p_title = '数学' + p_title  # '七年_________考试教学质量分析统计表（数学）'改成前面的了
    ws['D1'] = p_title
    ws['D1'].alignment = Alignment(horizontal='center')  # 设置标题居中
    ws['D1'].font = Font(size=20)  # 设置字号大小

    wb.save('汇总.xlsx')  # excel写入sheat


if __name__ == "__main__":
    main()
    print('all ok')
