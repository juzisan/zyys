# -*- coding: utf-8 -*-
#!/usr/bin/env python
"""
Created on Sat Nov  5 10:50:13 2016

@author: hello
"""


import os
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment



def duo(xlsx_nam):
    #文件名
    biao_or=pd.read_excel(xlsx_nam, index_col=0, skiprows = 1)#读excel
    biao = biao_or[-2:]
    lie_neirong = xlsx_nam.replace('试卷分析','')
    lie_neirong = re.search(r'(\d-\d)', xlsx_nam).group(1)
    biao['班级']= lie_neirong

    print ('duo done {}'.format(xlsx_nam))
    return biao



def main():
    #程序开始
    #数据区
    names = os.listdir(os.path.split(os.path.realpath(__file__))[0])
    shanchu = [i for i in names if i.startswith('汇总表')]
    for i in shanchu:
        os.remove(i)
    names = os.listdir(os.path.split(os.path.realpath(__file__))[0])
    names = [i for i in names if i.count('试卷分析')]
    print (names)
    #数据区
    data_hebing = [duo(i) for i in names]
    data_hebing = pd.concat(data_hebing, ignore_index=True)
    #data_hebing = data_hebing.set_index('班级')
    data_hebing = data_hebing.replace(np.nan, 0)
    #data_hebing = data_hebing.set_index('班级','姓名')
    biaotou = data_hebing.columns.values.tolist()

    data_hebing=data_hebing.pivot_table(index='班级', columns='姓名',  margins=False)

    biaotou = biaotou[1:-1]
    biaotou3 = []
    for i in biaotou:
        biaotou3.append((i,'扣分人数'))
        biaotou3.append((i,'平均扣分'))

    data_hebing = data_hebing[biaotou3]
    print (data_hebing)
    data_hebing.to_excel('汇总表表.xlsx')

    wb = load_workbook('汇总表表.xlsx')
    ws  = wb.active
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))#创建边框对象
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE#页面方向
    ws.page_setup.paperSize = ws.PAPERSIZE_A4#纸张大小
    ws.page_margins.left = 0.1
    ws.page_margins.rigt = 0
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0
    ws.page_margins.footer = 0#设置页边距
    ws.delete_rows(3)

    for row in ws.rows:
        for cell in row:
            cell.border =thin_border#设置边框

    ws.insert_rows(1)#插入标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)#合并标题行
    p_title = '九年_________考试教学质量分析统计表（数学）'
    ws['A1'] = p_title
    ws['A1'].alignment = Alignment(horizontal='center')#设置标题居中
    ws['A1'].font = Font(size=20)#设置字号大小
    ws['A3'] = '班级'

    ws.column_dimensions['A'].width = 2.6#设置列宽
    for i in range(2,ws.max_column+1):#最大列数目，是整数类型
        ws.column_dimensions[get_column_letter(i)].width = 3.4
    for i in range(4,ws.max_row+1):#最大列数目，是整数类型
        ws.row_dimensions[i].height = 40
    for row in list(ws.rows)[1:3]:#生成器转list才能迭代，前两行行
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)#设置文本自动换行

    lie_diyi = ws['A3:A'+str(ws.max_row)]
    for row in list(lie_diyi):#单元格选择区域是tuple，转换成list就能赋值了
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    for i in range(2,ws.max_column,2):#不能在插入第一行之前，否则保存的合并单元格有没有了，所以总是布局所有单元格的数据，然后再排版
        end_i = i +1
        ws.merge_cells(start_row=2, start_column=i, end_row=2, end_column=end_i)

    for row in list(ws.rows)[1:]:
        for cell in row:
            cell.border =thin_border#设置边框

    ziti = ws['B4:'+get_column_letter(ws.max_column)+str(ws.max_row)]
    for row in list(ziti):#单元格选择区域是tuple，转换成list就能赋值了
        for cell in row:
            cell.font=Font(size=7)
            cell.alignment = Alignment(wrap_text=True)#设置文本自动换行
    wb.save('汇总.xlsx')#excel写入sheat

    os.remove('汇总表表.xlsx')




if __name__ == "__main__":
    main()
    print ('all ok')

